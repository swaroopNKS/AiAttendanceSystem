from cgitb import text
import email
from time import sleep
from tkinter import *
from PIL import ImageTk, Image
import sqlite3
from importlib.resources import path
from msilib.schema import Directory
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
import xlsxwriter
import numpy as np
import pandas as pd
import face_recognition
import cv2
import os
import openpyxl


root = Tk()
user_id = StringVar()
passw = StringVar()

ref_label = Label(root, text="Admin Login").grid(columnspan=9, pady=10)
login_frame = Frame(root, border=5)
login_frame.grid(padx=50, pady=30)
root.title("Admin Login")
user_id_label =  Label(login_frame, text="User Id             :").grid(row=0,column=5, pady=11)
user_id_text_box = Entry(login_frame, text="Enter the User Id ", textvariable=user_id).grid(row=0, column=6, pady=11)
user_id_pass =  Label(login_frame, text="Password            :").grid(row=1,column=5)
user_id_pass_box = Entry(login_frame, text="Enter the Password :", textvariable=passw).grid(row=1, column=6)


conn = sqlite3.connect("admin_details.db")
cur = conn.cursor()
cur.execute("SELECT * FROM counts WHERE count_type = 'photo_id_ref'")
photo_id_ref_tup = cur.fetchall()
conn.commit()
print(len(photo_id_ref_tup))
if len(photo_id_ref_tup) == 0:
    cur.execute("INSERT INTO counts VALUES('photo_id_ref', 1)")
    conn.commit()

cur.execute("SELECT * FROM counts WHERE count_type = 'photo_id_ref'")
photo_id_ref_tup1 = cur.fetchall()
ref_list = photo_id_ref_tup1[0]
ref_tup_val = ref_list[1]
if ref_tup_val == 1:
    parent_directory = os.getcwd()
    directory_name = "stuORempImages"
    path = os.path.join(parent_directory, directory_name)
    os.mkdir(path)
    conn.execute("UPDATE counts SET count1 = 0 WHERE count_type = 'photo_id_ref'")

def attendance(name_id, time_taken):
    
    
    parent_dir = "F:/mini_project/attendence_sheets/"
    directory = str(date.today())
    path = os.path.join(parent_dir, directory) 
    isdir = os.path.isdir(path)
    if isdir == False:
        parent_dir = "F:/mini_project/attendence_sheets/"
        directory = str(date.today())
        path = os.path.join(parent_dir, directory) 
        os.mkdir(path)
        now2 = datetime.now()
        current_time2 = now2.strftime("%H:%M:%S")
        ref2 = str(current_time2)
        ref2_list = ref2.split(":")
        ref234 = ''.join(ref2_list)
        path_ref = "F:/mini_project/attendence_sheets/" + str(date.today()) +"/" + str(ref234) + ".xlsx"
        workbook = xlsxwriter.Workbook(path_ref)
        worksheet = workbook.add_worksheet() 
        worksheet.write(0,0, "Student Id")
        worksheet.write(0,1, "Student Name")
        worksheet.write(0,2, "Time")

        for i in range(1, len(name_id)+1):
            worksheet.write(i, 0, name_id[i - 1])

        for i in range(1, len(name_id)+1):
            conn = sqlite3.connect("admin_details.db")
            cur = conn.cursor()
            cur.execute("SELECT first_name FROM STUDENT_DETAILS WHERE id = '"+str(name_id[i - 1])+"';")
            ref = cur.fetchall()
            if len(ref) >= 1:
                ref1 = ref[0]
                if len(ref1) >= 0:
                    ref2 = ref1[0]
            worksheet.write(i, 1, ref2)

        for i in range(1, len(name_id)+1):
            worksheet.write(i, 2, time_taken[i - 1])

        workbook.close()

    else:
        
        now2 = datetime.now()
        current_time2 = now2.strftime("%H:%M:%S")
        ref2 = str(current_time2)
        ref2_list = ref2.split(":")
        ref234 = ''.join(ref2_list)
        path_ref = "F:/mini_project/attendence_sheets/" + str(date.today())+"/" + str(ref234) + ".xlsx"
        workbook = xlsxwriter.Workbook(path_ref)
        worksheet = workbook.add_worksheet() 
        worksheet.write(0,0, "Student Id")
        worksheet.write(0,1, "Student Name")
        worksheet.write(0,2, "Time")

        for i in range(1, len(name_id)+1):
            worksheet.write(i, 0, name_id[i - 1])

        for i in range(1, len(name_id)+1):
            conn = sqlite3.connect("admin_details.db")
            cur = conn.cursor()
            cur.execute("SELECT first_name FROM STUDENT_DETAILS WHERE id = '"+str(name_id[i - 1])+"';")
            ref = cur.fetchall()
            if len(ref) >= 1:
                ref1 = ref[0]
                if len(ref1) >= 0:
                    ref2 = ref1[0]
            worksheet.write(i, 1, ref2)

        for i in range(1, len(name_id)+1):
            worksheet.write(i, 2, time_taken[i - 1])
        
        workbook.close()

        



def markAttend():
    conn = sqlite3.connect("admin_details.db")
    cur = conn.cursor()
    
    







    parent_directory = os.getcwd()
    directory_name = "stuORempImages"
    path = os.path.join(parent_directory, directory_name)
    print(path)
    images = []
    personName = []
    myList = os.listdir(path)
    print(myList)
    name_id_list = []
    time_list = []
    for cu_img in myList:
        current_img = cv2.imread(f'{path}/{cu_img}')
        images.append(current_img)
        personName.append(os.path.splitext(cu_img)[0])
    print(personName)
    encodeListKnown = faceEncodings(images)
    print("All encodings Completed !!!")
    cap = cv2.VideoCapture(0)
    while True:
        ret, frame = cap.read()
        faces = cv2.resize(frame, (0, 0), None, 0.25, 0.25)
        faces = cv2.cvtColor(faces, cv2.COLOR_BGR2RGB)


        facesCurrentFrame = face_recognition.face_locations(faces)
        encodesCurrentFrame = face_recognition.face_encodings(faces, facesCurrentFrame)

        for encodeFace, faceLoc in zip(encodesCurrentFrame, facesCurrentFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = personName[matchIndex].upper()
                if name not in name_id_list:
                    name_id_list.append(name)
                    now1 = datetime.now()
                    current_time1 = now1.strftime("%H:%M:%S")
                    ref1 = str(current_time1)
                    time_list.append(ref1)
                y1, x2, y2, x1, = faceLoc
                y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(frame, (x1, y2-35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(frame, name, (x1 + 6, y2 + 6), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)
        cv2.imshow("Camera", frame)
        if cv2.waitKey(10) == 13:
            break
    cap.release()
    cv2.destroyAllWindows()
    attendance(name_id_list, time_list)
def faceEncodings(images):

    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)
        if len(encode) != 0:
            encodeList.append(encode[0])
    return encodeList



def takePhoto():
    cam = cv2.VideoCapture(0)
    
    img_counter = 0
    while True:
        ret, frame = cam.read()
        if not ret:
            print("failed to grab frame")
            break
        cv2.imshow("test", frame)
    
        k = cv2.waitKey(1)
        if k%256 == 27:
            # ESC pressed
            cv2.destroyAllWindows()
            print("Escape hit, closing...")
            break
        elif k%256 == 32:
            ref_admin_tupx = "SELECT * FROM counts WHERE count_type = 'student_user_id_count'"
            cur.execute(ref_admin_tupx)
            ref_admin_tup = cur.fetchall()
            conn.commit()
            if len(ref_admin_tup) >= 1:
                print(ref_admin_tup)
                ref_admin_tup1 = ref_admin_tup[0]
                ref_admin_id = ref_admin_tup1[1]
                y = ref_admin_id + 1
                img_name = "{}.png".format(y)
                parent_directory_ref = os.getcwd()
                directory_name_ref = "stuORempImages"
                path_ref = os.path.join(parent_directory_ref, directory_name_ref)
                cv2.imwrite(os.path.join(path_ref, img_name), frame)
                print(os.path.join(path_ref, img_name))
                print("{} written!".format(img_name))

def verify():
    ref_firstName = first_name_entry.get()
    ref_secondName = second_name_entry.get()
    ref_email = email_entry.get()
    ref_phone_num = phone_entry.get()
    ref_pass = password_entry.get()
    
    if ref_firstName == "" or ref_secondName == "" or ref_email == "" or ref_phone_num == "" or ref_pass == "":
        warn_label = Label(privilages, text="Please fill all the details!!!")
        warn_label.grid(row=7, column=1, columnspan=2, padx=5, pady=5)
    else:
        conn = sqlite3.connect("admin_details.db")
        cur = conn.cursor()
        ref_admin_tupx = "SELECT * FROM counts WHERE count_type = 'admin_user_id_count'"
        cur.execute(ref_admin_tupx)
        ref_admin_tup = cur.fetchall()
        conn.commit()
        if len(ref_admin_tup) >= 1:
            print(ref_admin_tup)
            ref_admin_tup1 = ref_admin_tup[0]
            ref_admin_id = ref_admin_tup1[1]
            x = ref_admin_id + 1
            ref_count_update = "UPDATE counts SET count1 = '"+str(x)+"' WHERE count_type = 'admin_user_id_count'"
            cur.execute(ref_count_update)
            conn.commit()
            ref_secondary_table_insert = "INSERT INTO ADMINS VALUES('"+str(ref_firstName)+"','" + str(ref_secondName)+"','"+ str(x) + "','"+ str(ref_pass) +"'," + str(ref_phone_num)+")"
            cur.execute(ref_secondary_table_insert)
            conn.commit()
            status_label1 = Label(privilages, text="User Id is : "+str(x))
            status_label1.grid(row=7, column=1, columnspan=2)
            status_label2 = Label(privilages, text="Password is : "+str(ref_pass))
            status_label2.grid(row=8, column=1, columnspan=2)


def adminPriv():
    global privilages, first_name_entry
    global second_name_entry
    global email_entry
    global phone_entry
    global password_entry
    global submit_button
    conn = sqlite3.connect("admin_details.db")
    cur = conn.cursor()
    #cur.execute("CREATE TABLE SECONDARY_ADMINS(first_name text, second_name text, email text, id text, password text, ph_num integer)")
    #conn.commit()
    #cur.execute("CREATE TABLE counts(count_type text, count1 integer)")
    #conn.commit()
    #cur.execute("INSERT INTO counts VALUES('admin_user_id_count', 9999)")
    #conn.commit()
    #cur.execute("SELECT * FROM counts")
    #print(cur.fetchall())
    #cur.execute("DELETE FROM counts")
    #conn.commit()
    privilages = Tk()
    privilages.title("Give previlages")

    first_name_label = Label(privilages, text = "First name : ")
    first_name_label.grid(row=1, column=1)
    first_name_entry = Entry(privilages, text="First Name")
    first_name_entry.grid(row=1, column=2, padx=11, pady=11)
    second_name_label = Label(privilages, text= "Second name").grid(row=2, column=1)
    second_name_entry = Entry(privilages, text="Second name")
    second_name_entry.grid(row=2, column=2, padx=11, pady=11)
    email_label = Label(privilages, text= " Email :").grid(row=3, column=1)
    email_entry = Entry(privilages, text="Email")
    email_entry.grid(row=3, column=2, padx=11, pady=11)
    phone_label = Label(privilages, text= " Mobile number :").grid(row=4, column=1)
    phone_entry = Entry(privilages,  text="Phone number")
    phone_entry.grid(row=4, column=2, padx=11, pady=11)
    password_label = Label(privilages, text= " Password :").grid(row=5, column=1)
    password_entry = Entry(privilages, text="Password")
    password_entry.grid(row=5, column=2, padx=11, pady=11)
    submit_button = Button(privilages, text="Submit", command=verify).grid(row=6, column=2)
    

def stuVerify():
    ref_stuFirstName = stu_first_name_entry.get()
    ref_stuSecondName = stu_second_name_entry.get()
    ref_stuEmail = stu_email_entry.get()
    ref_stuPhone_num = stu_phone_entry.get()
    ref_stuPass = stu_password_entry.get()
    
    if ref_stuFirstName == "" or ref_stuSecondName == "" or ref_stuEmail == "" or ref_stuPhone_num == "" or ref_stuPass == "":
        warn_label = Label(privilages, text="Please fill all the details!!!")
        warn_label.grid(row=7, column=1, columnspan=2, padx=5, pady=5)
    else:
        conn = sqlite3.connect("admin_details.db")
        cur = conn.cursor()
        ref_admin_tupx = "SELECT * FROM counts WHERE count_type = 'student_user_id_count'"
        cur.execute(ref_admin_tupx)
        ref_admin_tup = cur.fetchall()
        conn.commit()
        if len(ref_admin_tup) >= 1:
            print(ref_admin_tup)
            ref_admin_tup1 = ref_admin_tup[0]
            ref_admin_id = ref_admin_tup1[1]
            y = ref_admin_id + 1
            ref_count_update = "UPDATE counts SET count1 = '"+str(y)+"' WHERE count_type = 'student_user_id_count'"
            cur.execute(ref_count_update)
            conn.commit()
            ref_secondary_table_insert = "INSERT INTO STUDENT_DETAILS VALUES('"+str(ref_stuFirstName)+"','" + str(ref_stuSecondName)+"','" + str(ref_stuEmail) +"','"+ str(y) + "','"+ str(ref_stuPass) +"'," + str(ref_stuPhone_num)+")"
            cur.execute(ref_secondary_table_insert)
            conn.commit()
            status_label1 = Label(add_stu_window, text="User Id is : "+str(y))
            status_label1.grid(row=7, column=1, columnspan=2)
            status_label2 = Label(add_stu_window, text="Password is : "+str(ref_stuPass))
            status_label2.grid(row=8, column=1, columnspan=2)


def addStudents():
    global add_stu_window, stu_first_name_entry
    global stu_second_name_entry
    global stu_email_entry
    global stu_phone_entry
    global stu_password_entry
    global stu_submit_button
    conn = sqlite3.connect("admin_details.db")
    cur = conn.cursor()
    #cur.execute("CREATE TABLE STUDENT_DETAILS(first_name text, second_name text, email text, id text, password text, ph_num integer)")
    #conn.commit()
    #cur.execute("CREATE TABLE counts(count_type text, count1 integer)")
    #conn.commit()
    #cur.execute("INSERT INTO counts VALUES('student_user_id_count', 9999)")
    #conn.commit()
    #cur.execute("SELECT * FROM counts")
    #print(cur.fetchall())
    #cur.execute("DELETE FROM counts")
    #conn.commit()
    cur.execute("SELECT * FROM STUDENT_DETAILS")
    print(cur.fetchall())
    add_stu_window = Tk()
    add_stu_window.title("Add Students")

    stu_first_name_label = Label(add_stu_window, text = "First name : ")
    stu_first_name_label.grid(row=1, column=1)
    stu_first_name_entry = Entry(add_stu_window, text="First Name")
    stu_first_name_entry.grid(row=1, column=2, padx=11, pady=11)
    stu_second_name_label = Label(add_stu_window, text= "Second name").grid(row=2, column=1)
    stu_second_name_entry = Entry(add_stu_window, text="Second name")
    stu_second_name_entry.grid(row=2, column=2, padx=11, pady=11)
    stu_email_label = Label(add_stu_window, text= " Email :").grid(row=3, column=1)
    stu_email_entry = Entry(add_stu_window, text="Email")
    stu_email_entry.grid(row=3, column=2, padx=11, pady=11)
    stu_phone_label = Label(add_stu_window, text= " Mobile number :").grid(row=4, column=1)
    stu_phone_entry = Entry(add_stu_window,  text="Phone number")
    stu_phone_entry.grid(row=4, column=2, padx=11, pady=11)
    stu_password_label = Label(add_stu_window, text= " Password :").grid(row=5, column=1)
    stu_password_entry = Entry(add_stu_window, text="Password")
    stu_password_entry.grid(row=5, column=2, padx=11, pady=11)
    stu_takePhoto_entry = Button(add_stu_window, text="Take a photo", command=takePhoto).grid(row=6, column=2)
    stu_submit_button = Button(add_stu_window, text="Submit", command=stuVerify).grid(row=7, column=2)
    
def login():
    global attendence_window
    global bg_img
    ref_user_id = user_id.get()
    print(ref_user_id)
    ref_pass = passw.get()
    conn = sqlite3.connect("admin_details.db")
    cur = conn.cursor()
    ref_str = "SELECT * FROM ADMINS WHERE id = '"+str(ref_user_id)+"'"
    #cur.execute(ref_str)
    #print(cur.fetchall())
    cur.execute(ref_str)
    ref_tup = cur.fetchall()
    print(ref_user_id)
    print(ref_tup)
    if len(ref_tup) >= 1:
        if ref_tup[0][2] == ref_user_id :
            if ref_tup[0][3] == ref_pass:
                Label(login_frame, text="Login  successfull !!!!!").grid(row=3, column=5,columnspan=2)
                Label(login_frame,text="Wait .....").grid(row=4, column=5,columnspan=2)
                root.destroy()
                attendence_window = Tk()
                attendence_window.title("Main Navigation")
                head_label = Label(attendence_window, text="Main Navigations", font="italic")
                head_label.grid(row=0, column=0, columnspan=10, padx=11, pady=9)
                bg_img = ImageTk.PhotoImage(Image.open("Images\main_bg1.jpg"))
                bg_img_label = Label(attendence_window, image=bg_img)
                bg_img_label.grid(row=1, column=0,padx=18, pady=18, ipadx=11, ipady=11, columnspan=5)
                add_admin_priv_but = Button(attendence_window, text="Give Admin privilages", fg="white", bg="blue", command=lambda : adminPriv())
                add_admin_priv_but.grid(row=2, column=0, padx=18, pady=18, ipadx=3 )
                mark_att_but = Button(attendence_window, text="Mark Attendence", fg="white", bg="blue", command=markAttend)
                mark_att_but.grid(row=2, column=2, padx=18, pady=18, ipadx=3 )
                add_stu_but = Button(attendence_window, text="Add Students", fg="white", bg="blue", command=addStudents)
                add_stu_but.grid(row=2, column=4, padx=18, pady=18, ipadx=3)

            else:
                Label(login_frame, text="Login unsuccessfull !!!").grid(row=3, column=5,columnspan=2)
        else:    
            Label(login_frame, text="Login unsuccessfull !!!").grid(row=3, column=5,columnspan=2)
    else:
        Label(login_frame, text="Login unsuccessfull !!!").grid(row=3, column=5,columnspan=2)


login_submit = Button(login_frame, text="Submit", bg="blue", fg="white", command=login).grid(row=2, column=5, columnspan=2, pady=18)
#conn = sqlite3.connect("admin_details.db")
#cur = conn.cursor()
#cur.execute("INSERT INTO ADMINS VALUES('Admin', 'Admin', '99999', 'Admin', '9999999999')")
#conn.commit()



"""
count = 0
if count == 0:
    conn = sqlite3.connect("admin_details.db")
    cur = conn.cursor()
    #cur.execute(""" 
    #  CREATE TABLE ADMINS(
    #    first_name text,
    #    last_name text,
    #    id text,
    #    password text,
    #    address text,
    #    pn_number integer
    #)
    # """)
    
"""  x = "INSERT INTO ADMINS VALUES('Admin', 'Admin', 'Admin', 'Admin', 'Admin', '1234567890')"
    cur.execute(x)
    conn.commit()

    
    #cur.execute("DELETE FROM ADMINS")
    #conn.commit()
    x = 'Admin'
    cur.execute("SELECT * FROM ADMINS WHERE  first_name = 'Admin'")
    print(cur.fetchall())
    conn.commit()
    count = 1
"""

conn = sqlite3.connect("admin_details.db")
cur = conn.cursor()
#cur.execute("CREATE TABLE STUDENT_DETAILS(first_name text, second_name text, email text, id text, password text, ph_num integer)")
#conn.commit()
#cur.execute("CREATE TABLE counts(count_type text, count1 integer)")
    #conn.commit()
    #cur.execute("INSERT INTO counts VALUES('student_user_id_count', 9999)")
    #conn.commit()
cur.execute("SELECT * FROM STUDENT_DETAILS")
print("hello", cur.fetchall())












conn = sqlite3.connect("admin_details.db")
cur = conn.cursor()
cur.execute("SELECT first_name FROM STUDENT_DETAILS WHERE id = '"+str(100000)+"';")
ref = cur.fetchall()
if len(ref) >= 1:
    ref1 = ref[0]
    if len(ref1) >= 0:
        ref2 = ref1[0]
        print(ref2)





root.mainloop()